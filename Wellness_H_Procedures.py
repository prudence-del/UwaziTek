import pandas as pd
import pdfplumber
import tkinter as tk
from tkinter import filedialog
from datetime import datetime, timedelta
import re
from openpyxl.workbook import Workbook
from openpyxl.styles import PatternFill
from fuzzywuzzy import process
import json

#%% Base data setup
# loading data with services offered in the hospital and their cost
Procedure_file_path = 'synthea_sample_data_csv_latest/procedures.csv'
Medication_file_path = 'synthea_sample_data_csv_latest/medications.csv'
Base_data = pd.read_csv(Procedure_file_path)
medication_data = pd.read_csv(Medication_file_path)
# reading the first 5 rows of the data
print(Base_data.head(5))
# detailed info about the data
print(Base_data.info())

#%% data cleaning
# procedure base data cleaning
# dropping unnecessary columns
Base_data = Base_data.drop(
    ['START', 'STOP', 'PATIENT', 'ENCOUNTER',
     'SYSTEM', 'CODE', 'REASONCODE', 'REASONDESCRIPTION'], axis=1)  # columns
# columns left
print(Base_data.columns)
print(Base_data.head(5))

# drop duplicates (if service and cost are identical)
# keep the first occurrence
Base_data = Base_data.drop_duplicates(subset=['DESCRIPTION', 'BASE_COST'], keep='first')

# max value of different charges but same service
# as_index=false will group description as regular column to avoid fixed(indexing)
Base_data = Base_data.groupby('DESCRIPTION', as_index=False).agg({'BASE_COST': 'max'})

# format base_cost column to 2 decimal places
Base_data['BASE_COST'] = Base_data['BASE_COST'].map(lambda x: f"{x:.2f}")
# renaming columns {'OldName':'NewName'}
Base_data = Base_data.rename(columns={'DESCRIPTION': 'Services provided at Wellness Hospital'})

# Save the grouped(base) data to an Excel file in the current working directory
output_file_name = 'Base_data_report.xlsx'
Base_data.to_excel(output_file_name, index=False)
print(f"Report successfully saved to the current directory as {output_file_name}\n")

# medication base data cleaning
medication_data = medication_data.drop(
    ['START', 'STOP', 'PATIENT', 'PAYER', 'ENCOUNTER', 'CODE', 'PAYER_COVERAGE',
     'DISPENSES', 'TOTALCOST', 'REASONCODE', 'REASONDESCRIPTION'], axis=1)
# printing the first 5 rows after dropping the columns
print(medication_data.head(5))

# convert the description column to lower case to avoid retaining same medications
medication_data['DESCRIPTION'] = medication_data['DESCRIPTION'].str.lower()

# dropping duplicates
# count of the initial rows
initial_rows = len(medication_data)

# remaining rows after duplicates dropped
final_rows = medication_data.drop_duplicates(subset=['DESCRIPTION', 'BASE_COST'], keep='first')
final_rows = final_rows.shape[0]  # number of rows

# duplicate rows
dropped_rows = initial_rows - final_rows
print(f"dropped duplicates rows: {dropped_rows}")

# remaining rows (will act as a base medication data)
print(f"remaining rows: {final_rows}")

# renaming columns
medication_data = medication_data.rename(columns={'DESCRIPTION': 'Wellness Medication'})
print(medication_data.columns)

# max value of same medication names with different cost
medication_data = medication_data.groupby('Wellness Medication', as_index=False).agg({'BASE_COST': 'max'})

# converting cost column to 2 decimal places
medication_data['BASE_COST'] = medication_data['BASE_COST'].map(lambda x: f"{x:.2f}")

# saving medication cleaned data as an Excel file
report_file_name = 'medication_base_report.xlsx'
medication_data.to_excel(report_file_name, index=False)
print(f"report saved successfully: {report_file_name}")


#%% Invoice upload, text extract, image extract, convert text to data frame
# invoice upload
def upload_file():
    root = tk.Tk()  # initializes the root window
    root.withdraw()  # pop-up only of the dialog box
    invoice_file_path = filedialog.askopenfilename(title="Select invoice")
    return invoice_file_path



# Function to extract and save text from PDF

# Function to extract text from the PDF
def extract_text_from_pdf(pdf_invoice_path):
    try:
        # Open the PDF with pdfplumber
        with pdfplumber.open(pdf_invoice_path) as pdf:
            # Extract text from all pages of the PDF
            invoice_text = "\n".join([page.extract_text() or "" for page in pdf.pages if page.extract_text()])

            # Display the extracted text
            if invoice_text:
                print("Extracted Invoice Text:\n")
                print(invoice_text)
                return invoice_text  # return the extracted text

            else:
                print("No text found in the invoice.")
                return ""  # return empty string if no text found
    except Exception as e:
        print(f"Error extracting text from PDF: {e}")
        return ""


def clean_invoice_text(invoice_text):
    """
    Cleans the invoice text by replacing line breaks after field names.
    Ensures all field values stay on a single line.
    """
    # Replace line breaks between keys and values
    cleaned_text = re.sub(r'(\w+:)\s*\n', r'\1 ', invoice_text)
    return cleaned_text
# check for serial number (a must in invoices)

def check_serial_number(invoice_text, expected_serial_prefix="Serial number: #"):
    """
    Checks if the serial number exists in the invoice text and matches the expected format.

    Returns:
        bool: True if the serial number is found and valid, False otherwise.
    """
    # Regular expression to match "Serial number: #<digits>"
    serial_pattern = rf"{re.escape(expected_serial_prefix)}(\d+)"
    match = re.search(serial_pattern, invoice_text)

    if match:
        serial_number = match.group(1)
        print(f"Serial number found: {serial_number}")
        return True
    else:
        print("Serial number not found or invalid.")
        return False


#%%  validation of the invoices
# check before the fraud test
# mandatory fields
def check_mandatory_fields(invoice_text):
    """
    Check for mandatory fields in the invoice text.
    Returns metadata and reasons for rejection.
    """
    reasons = []
    mandatory_fields = ['Serial number', 'Patient Name', 'Invoice No', 'Invoice Date', 'Bill to', 'Bank Name', 'Bank Account']
    metadata = {}

    # Define regex patterns for the mandatory fields
    patterns = {
        'Serial number': r'Serial\s+number:\s*#?(\d+)',
        'Patient Name': r'Patient Name:\s*([A-Za-z\s]+(?:\s+[A-Za-z\s]+)*)',
        'Invoice No': r'Invoice\s*N0?:\s*(WELLNESS-INV/\d{10})',
        'Invoice Date': r'Invoice Date:\s*([\d]{2}\s[A-Za-z]+,\s[\d]{4})',
        'Bank Name': r'Bank Name:\s*([\w\s,&]+)',
        'Bill to': r'Bill to:\s*([\w\s]+)',
        'Bank Account': r'Bank Account:\s*(\d{4}\s\d{4}\s\d{4})'
    }
    cleaned_text = clean_invoice_text(invoice_text)
    # Check for each mandatory field
    for field in mandatory_fields:
        pattern = patterns.get(field, None)
        if pattern:
            match = re.search(pattern, cleaned_text)
            if match:
                metadata[field] = match.group(1)
            else:
                metadata[field] = None
                reasons.append(f"Missing mandatory field: {field}")

        # Specific check for Invoice Date (e.g., future dates or outdated invoices)
        if field == "Invoice Date" and metadata.get("Invoice Date"):
            try:
                invoice_date = datetime.strptime(metadata["Invoice Date"], "%d %B, %Y")
                if invoice_date > datetime.now():
                    reasons.append("Invoice date cannot be in the future")
                elif invoice_date < datetime.now() - timedelta(days=90):
                    reasons.append("Invoice date is older than 3 months")
            except ValueError:
                reasons.append("Invalid date format (expected format: 'DD Month, YYYY')")

    # Ensure Hospital Name is added to metadata if not present
    if "Hospital Name" not in metadata:
        metadata["Hospital Name"] = "Wellness Hospital"

    return metadata, reasons


# Function to save rejection reasons to JSON
def save_reasons_to_json(reasons, filename="invoice_rejection_reasons.json"):
    """
    Save the rejection reasons to a JSON file.
    """
    data = {"reasons": reasons}
    try:
        with open(filename, "w") as json_file:
            json.dump(data, json_file, indent=4)
            print(f"Rejection reasons successfully saved to {filename}")
    except Exception as e:
        print(f"Error saving rejection reasons to JSON: {e}")



def extract_invoice_items(invoice_text):

    item_pattern = r"(\d{2}/\d{2}/\d{4})\s+([A-Za-z0-9\s&/()-]+)\s+\d+\.\d+\s+\d+\.\d+\s+(\d+\.\d+)"

    items = []
    matches = re.findall(item_pattern, invoice_text)

    for match in matches:
        date = match[0]
        description = match[1].strip()  # Extract the description
        amount = float(match[2].replace(',', ''))  # Extract and convert the amount to float

        items.append({'DATE': date, 'DESCRIPTION': description, 'AMOUNT': amount})

    # Create a DataFrame from the extracted items
    if items:  # Only create a DataFrame if items are found
        items_df = pd.DataFrame(items)
    else:
        print("No invoice items found. Please check the regex or invoice format.")
        items_df = pd.DataFrame(
            columns=['DATE', 'DESCRIPTION', 'AMOUNT'])

    print("Extracted Invoice Items:\n", items_df)
    return items_df


# json file to be fetched by app-user for services confirmations
def save_to_json(metadata, items, filename="invoice_data.json"):
    data = {
        "metadata": metadata,
        "items": items.to_dict(orient="records")  # Convert DataFrame to list of dictionaries
    }
    try:
        with open(filename, "w") as json_file:
            json.dump(data, json_file, indent=4)
            print(f"Data successfully saved to {filename}")
    except Exception as e:
        print(f"Error saving data to JSON: {e}")


# Process invoice for fraud detection
def process_invoice(pdf_file_path):
    """
    Processes the invoice by checking mandatory fields, validating content,
    and performing fraud detection if checks pass.
    """
    # Step 1: Extract text from the PDF
    print("Extracting text from invoice...")
    invoice_text = extract_text_from_pdf(pdf_file_path)

    if not invoice_text:
        rejection_reasons = ["No text found in the invoice."]
        save_reasons_to_json(rejection_reasons)  # Save the reason to JSON
        print("No text found. Exiting process.")
        return  # Stop processing by returning nothing

    # Step 2: Check for Serial Number first (critical field)
    print("Checking for serial number...")
    if not check_serial_number(invoice_text):
        rejection_reasons = ["Serial number not found. Invoice cannot be processed."]
        save_reasons_to_json(rejection_reasons)  # Save the reason to JSON
        print("Serial number missing. Exiting process.")
        return  # Stop processing

    # Step 3: Check for all mandatory fields
    print("Checking for mandatory fields...")
    metadata, mandatory_reasons = check_mandatory_fields(invoice_text)

    if mandatory_reasons:  # If any mandatory field is missing or invalid
        print("Mandatory field validation failed. Reasons:")
        for reason in mandatory_reasons:
            print(f"- {reason}")
        save_reasons_to_json(mandatory_reasons)  # Save reasons to JSON
        print("Mandatory fields incomplete. Exiting process.")
        return  # Stop processing

    # Step 4: Save valid extracted data to a JSON file
    print(f"Extracted Metadata: {metadata}")

    try:
        # Extract invoice items (description and amounts)
        print("Extracting invoice items...")
        items_df = extract_invoice_items(invoice_text)

        # Save invoice data to JSON file
        invoice_data = {
            'metadata': metadata,
            'items': items_df.to_dict(orient='records')  # Convert DataFrame to list of dictionaries
        }
        json_file_path = "invoice_data.json"

        with open(json_file_path, 'w') as json_file:
            json.dump(invoice_data, json_file, indent=4)
        print(f"Invoice data saved successfully as {json_file_path}")
    except Exception as e:
        print(f"Error saving invoice data: {e}")
        return  # Exit if saving data fails

    # Step 5: Proceed to fraud detection
    print("Invoice validated successfully. Proceeding with fraud detection...")


    if len(fraud_results) > 0:
        generate_combined_report(fraud_results, metadata)  # Generate fraud report
    else:
        print("No fraud results to report. Skipping report generation.")

#%% data normalization for easy comparison of invoice and base data
#  after cleaning
base_data = pd.read_excel("Base_data_report.xlsx")
medication_data = pd.read_excel("medication_base_report.xlsx")


# for normalizing text to ensure consistence in formating
def normalize_description(desc):
    return re.sub(r"\s+", " ", desc.lower().strip())


base_data['Normalized DESCRIPTION'] = base_data['Services provided at Wellness Hospital'].apply(normalize_description)
medication_data['Normalized DESCRIPTION'] = medication_data['Wellness Medication'].apply(normalize_description)


#%% comparison of the invoice items and their prices with that of the base cost
# fraud detection and categorization
# different colors depending on the extent of fraud.

def compare_invoice_with_base(invoice_items, hospital_base_data, medication_base_data):
    # Print base descriptions once for verification
    print("Base data item descriptions (hospital):")
    print(hospital_base_data['Normalized DESCRIPTION'].head())

    print("Base data item descriptions (medication):")
    print(medication_base_data['Normalized DESCRIPTION'].head())
    # Normalize the DESCRIPTION column for both datasets if not already done
    if 'Normalized DESCRIPTION' not in hospital_base_data.columns:
        hospital_base_data['Normalized DESCRIPTION'] = hospital_base_data['DESCRIPTION'].apply(normalize_description)

    if 'Normalized DESCRIPTION' not in medication_base_data.columns:
        medication_base_data['Normalized DESCRIPTION'] = medication_base_data['DESCRIPTION'].apply(normalize_description)

    comparison_results = []
    unmatched_items = []
    total_invoice_amount = invoice_items['AMOUNT'].sum()  # Calculate total invoice amount

    for _, item in invoice_items.iterrows():
        description = normalize_description(item['DESCRIPTION'])
        invoice_cost = item['AMOUNT']

        # Check if the item is a medication or a procedure
        if 'medication' in description.lower():
            base_data = medication_base_data
            is_medication = True
        else:
            base_data = hospital_base_data
            is_medication = False


        base_descriptions = base_data['Normalized DESCRIPTION'].tolist()
        match = process.extractOne(description, base_descriptions)

        if match:
            base_row = base_data[base_data['Normalized DESCRIPTION'] == match[0]]
        else:
            base_row = pd.DataFrame()

        # If no match is found
        if base_row.empty:
            unmatched_items.append(description)
            comparison_results.append({
                'Description': item['DESCRIPTION'],
                'Invoice Cost': invoice_cost,
                'Base Cost': None,
                'Price Difference': None,
                'Fraud Category': "Service not found in base data"
            })
            continue

        # Base cost and fraud categorization logic
        base_cost = float(base_row['BASE_COST'].values[0])
        price_difference = round(invoice_cost - base_cost, 2)

        if is_medication:
            if invoice_cost == base_cost or (invoice_cost - base_cost) <= 20:
                fraud_category = "Legitimate"
            elif invoice_cost > base_cost:
                fraud_category = "Risk" if (invoice_cost - base_cost) <= 200 else "Fraud"
            else:
                fraud_category = "Potential Underreporting"
        else:
            if invoice_cost == base_cost or (invoice_cost - base_cost) <= 100:
                fraud_category = "Legitimate"
            elif invoice_cost > base_cost:
                fraud_category = "Risk" if (invoice_cost - base_cost) <= 1000 else "Fraud"
            else:
                fraud_category = "Potential Underreporting"

        # Append result
        comparison_results.append({
            'Description': item['DESCRIPTION'],
            'Invoice Cost': invoice_cost,
            'Base Cost': base_cost,
            'Price Difference': price_difference,
            'Fraud Category': fraud_category
        })

    # overall fraud status of the invoice
    fraud_categories = [result['Fraud Category'] for result in comparison_results if result['Fraud Category']]
    if 'Fraud' in fraud_categories:
        overall_status = 'Fraud'
    elif fraud_categories.count('Risk') > 1:
        overall_status = 'Risky'
    elif all(category == 'Legitimate' for category in fraud_categories if category != "Service not found in base data"):
        overall_status = 'Legitimate'
    else:
        overall_status = 'Unknown'

    # Add summary rows
    total_row = {
        'Description': 'Total Invoice Amount',
        'Invoice Cost': total_invoice_amount,
        'Base Cost': None,
        'Price Difference': None,
        'Fraud Category': None
    }
    overall_status_row = {
        'Description': 'Overall Status',
        'Invoice Cost': None,
        'Base Cost': None,
        'Price Difference': None,
        'Fraud Category': overall_status
    }
    comparison_results.append(total_row)
    comparison_results.append(overall_status_row)

    # Invoice Description and invoice cost(to be compared with the base cost)
    print("\nSimplified Comparison Results:")
    for result in comparison_results:
        print(f"Description: {result['Description']}, Invoice Cost: {result['Invoice Cost']}")

    # Print unmatched items
    if unmatched_items:
        print("\nUnmatched invoice items (descriptions not found in base data):")
        for item in unmatched_items:
            print(f"- {item}")

    return comparison_results


def generate_combined_report(fraud_results, metadata):
    """
        Generates a fraud detection report with metadata and fraud categories.

        Parameters:
            fraud_results (list): A list of dictionaries with fraud detection details.
            metadata (dict): A dictionary containing metadata about the invoice.
        """

    # Clean metadata
    metadata = {key: value.replace("\n", " ") if isinstance(value, str) else value for key, value in
                metadata.items()}

    # Validate and clean fraud_results
    if not isinstance(fraud_results, list):
        print("Fraud results are not in a list format.")
        return
    fraud_results = [result for result in fraud_results if isinstance(result, dict)]

    # Check for valid data
    if not fraud_results:
        print("No valid fraud detection results to process.")
        return

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active  # active worksheet
    ws.title = "Fraud Detection Report"

    # Add metadata
    metadata_start_row = 1
    metadata_fields = [
        ("Hospital Name", metadata.get("Hospital Name", "N/A")),
        ("Bank Name", metadata.get("Bank Name", "N/A")),
        ("Bank Account", metadata.get("Bank Account", "N/A")),
        ("Patient Name", metadata.get("Patient Name", "N/A")),
        ("Invoice Number", metadata.get("Invoice No", "N/A")),
    ]

    for idx, (label, value) in enumerate(metadata_fields):
        ws[f"A{metadata_start_row + idx}"] = f"{label}:"
        ws[f"B{metadata_start_row + idx}"] = value

    # Add fraud results
    metadata_end_row = metadata_start_row + len(metadata_fields) + 1
    headers = ["Description", "Invoice Cost", "Base Cost", "Price Difference", "Fraud Category"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=metadata_end_row, column=col_idx, value=header)

    # Define color coding
    colors = {
        "Legitimate": "00FF00",  # green
        "Risk": "FFA500",  # orange
        "Fraud": "FF0000",  # red
        "Service not found in base data": "FF FF00"  # bright yellow
    }

    # Populate fraud results
    for row_idx, result in enumerate(fraud_results, start=metadata_end_row + 1):
        ws.cell(row=row_idx, column=1, value=result.get("Description", "N/A"))
        ws.cell(row=row_idx, column=2, value=result.get("Invoice Cost", "N/A"))
        ws.cell(row=row_idx, column=3, value=result.get("Base Cost", "N/A"))
        ws.cell(row=row_idx, column=4, value=result.get("Price Difference", "N/A"))
        fraud_category = result.get("Fraud Category", "Unknown")
        ws.cell(row=row_idx, column=5, value=fraud_category)

        # Apply color
        fill_color = colors.get(fraud_category, None)
        if fill_color:
            for col_idx in range(1, ws.max_column + 1):  # Loop through all columns in the row
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(start_color=fill_color,
                                                                        end_color=fill_color, fill_type="solid")

    # Save workbook
    report_file_path = "fraud_report_with_metadata_and_categories.xlsx"
    wb.save(report_file_path)
    print(f"Fraud detection report saved successfully as '{report_file_path}'!")

    # Save as JSON
    combined_report = {
        "Metadata": metadata,
        "Fraud Detection Results": fraud_results,
    }
    json_file_path = "fraud_report_with_metadata_and_categories.json"
    with open(json_file_path, "w") as json_file:
        json.dump(combined_report, json_file, indent=4)
    print(f"Fraud detection report saved successfully as '{json_file_path}'!")

    return report_file_path, json_file_path


def save_report(metadata, fraud_results, json_file="combined_report.json"):
    # Combine Metadata and Fraud Results
    combined_report = {
        "Metadata": metadata,
        "Fraud Detection Results": fraud_results,
    }

    # Save as JSON
    with open(json_file, "w") as json_out:
        json.dump(combined_report, json_out, indent=4)
    print(f"Report saved successfully as {json_file}")


pdf_file_path = upload_file()
invoice_text = extract_text_from_pdf(pdf_file_path)
invoice_items = extract_invoice_items(invoice_text)
fraud_results = compare_invoice_with_base(invoice_items, base_data, medication_data)
process_invoice(pdf_file_path)
